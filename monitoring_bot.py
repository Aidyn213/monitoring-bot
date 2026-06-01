"""
Magnum Monitoring Bot — облачная версия
Пользователь присылает xlsx → бот анализирует → отвечает отчётом + Excel
"""
 
import os
import io
import statistics
import telebot
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
 
BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
bot = telebot.TeleBot(BOT_TOKEN)
 
# ── Загрузка маппинга брендов ────────────────────────────────────────────────
def load_brands():
    brands = {}
    try:
        wb = load_workbook("Маппинг.xlsx", data_only=True)
        for row in wb['Бренды'].iter_rows(min_row=2, max_row=wb['Бренды'].max_row, values_only=True):
            try:
                if row[0] and (len(row) < 2 or row[1] != 'Закрыто'):
                    km    = row[1] if len(row) > 1 else '—'
                    brand = row[4] if len(row) > 4 else '—'
                    brands[str(row[0]).strip().upper()] = {
                        'km': km or '—', 'brand': brand or '—'}
            except (IndexError, TypeError):
                continue
    except Exception:
        pass
    return brands
 
BRANDS = load_brands()
 
def lookup(name):
    key = name.strip().upper()
    if key in BRANDS: return BRANDS[key]
    for k, v in BRANDS.items():
        if key in k or k in key: return v
    return {'km': '—', 'brand': '—'}
 
# ── Анализ ───────────────────────────────────────────────────────────────────
def run_analysis(file_bytes):
    wb_src = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws_src = wb_src['Тепловая карта']
 
    MARKET_IDXS = [4, 5, 6, 8]
    SMALL_IDXS  = [7, 10]
    KASPI_IDX   = 2
 
    def g(row, i):
        try: return row[i]
        except (IndexError, TypeError): return None
 
    segment, rows = None, []
    for row in ws_src.iter_rows(min_row=8, max_row=65, values_only=True):
        name = g(row, 0)
        if name in ('HARD', 'SOFT', 'СЕЗОН'):
            segment = name; continue
        if not name or name == '(пусто)': continue
 
        def num(val):
            return val if isinstance(val, (int, float)) else None
 
        magnum = num(g(row, 1))
        kaspi  = num(g(row, KASPI_IDX))
        market = [num(g(row, i)) for i in MARKET_IDXS if isinstance(g(row, i), (int, float))]
        smalls = [num(g(row, i)) for i in SMALL_IDXS  if isinstance(g(row, i), (int, float))]
        if smalls: market.append(round(statistics.mean(smalls)))
        if not market: continue
 
        avg  = round(statistics.mean(market))
        mn   = min(market)
        mx   = max(market)
        info = lookup(name)
 
        if magnum:
            pct = round((magnum - avg) / avg * 100, 1)
            if pct >= 15:    status = 'Вне рынка'
            elif pct >= 0:   status = 'Дороже'
            elif pct >= -5:  status = 'Норма'
            else:            status = 'Дешевле'
        else:
            pct, status = None, 'Нет цены'
 
        rows.append(dict(segment=segment, name=name, magnum=magnum, kaspi=kaspi,
                         avg=avg, min=mn, max=mx, pct=pct, status=status,
                         km=info['km'], brand=info['brand']))
    return rows
 
# ── Генерация Excel ──────────────────────────────────────────────────────────
def build_excel(rows):
    FN = 'Arial'
    HDR_FILL = PatternFill('solid', start_color='1F3864')
    HDR_FONT = Font(name=FN, bold=True, color='FFFFFF', size=10)
    SEG_FILLS = {'HARD':  PatternFill('solid', start_color='C6EFCE'),
                 'SOFT':  PatternFill('solid', start_color='DDEBF7'),
                 'СЕЗОН': PatternFill('solid', start_color='FFF2CC')}
    SF = {'Вне рынка':     PatternFill('solid', start_color='FFC7CE'),
          'Дороже':   PatternFill('solid', start_color='FFEB9C'),
          'Норма':    PatternFill('solid', start_color='FFFFFF'),
          'Дешевле':  PatternFill('solid', start_color='C6EFCE'),
          'Нет цены': PatternFill('solid', start_color='EDEDED')}
    SFONTS = {'Вне рынка':     Font(name=FN, color='9C0006', bold=True, size=10),
              'Дороже':   Font(name=FN, color='9C6500', size=10),
              'Норма':    Font(name=FN, size=10),
              'Дешевле':  Font(name=FN, color='276221', size=10),
              'Нет цены': Font(name=FN, color='7F7F7F', size=10)}
    NORM = Font(name=FN, size=10)
    BOLD = Font(name=FN, bold=True, size=10)
    thin = Side(style='thin', color='BFBFBF')
    BRD  = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    def cs(ws, r, c, val, fill=None, font=None, align='left', num_fmt=None, bold=False):
        cell = ws.cell(row=r, column=c, value=val)
        if fill:   cell.fill = fill
        cell.font      = font or (BOLD if bold else NORM)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border    = BRD
        if num_fmt: cell.number_format = num_fmt
 
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = 'Тепловая карта'
    ws1.freeze_panes = 'B3'
    date_str = datetime.now().strftime('%d.%m.%Y')
 
    HEADERS = ['Товар','Сегмент','Магнум','Ср.рынок','Откл, %','Статус','Каспий','Мин','КМ','Бренд']
    WIDTHS  = [42, 8, 9, 9, 10, 11, 9, 9, 22, 18]
 
    ws1.merge_cells('A1:J1')
    t = ws1['A1']
    t.value     = f'Мониторинг цен vs конкуренты — {date_str}'
    t.font      = Font(name=FN, bold=True, size=13, color='1F3864')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 22
 
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cs(ws1, 2, ci, h, fill=HDR_FILL, font=HDR_FONT, align='center')
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[2].height = 18
 
    r, prev_seg = 3, None
    for row in rows:
        seg, status = row['segment'], row['status']
        if seg != prev_seg:
            ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            sc = ws1.cell(row=r, column=1, value=f'  {seg}')
            sc.fill = SEG_FILLS[seg]
            sc.font = Font(name=FN, bold=True, size=10)
            sc.alignment = Alignment(horizontal='left', vertical='center')
            sc.border = BRD
            ws1.row_dimensions[r].height = 16
            r += 1
            prev_seg = seg
 
        pct_val = (row['pct'] / 100) if row['pct'] is not None else None
        cs(ws1, r, 1,  row['name'],   font=SFONTS[status])
        cs(ws1, r, 2,  seg,           align='center')
        cs(ws1, r, 3,  row['magnum'], fill=SF[status], font=SFONTS[status], align='right', num_fmt='#,##0')
        cs(ws1, r, 4,  row['avg'],    align='right', num_fmt='#,##0')
        cs(ws1, r, 5,  pct_val,       fill=SF[status], font=SFONTS[status], align='right', num_fmt='+0.0%;-0.0%;"-"')
        cs(ws1, r, 6,  status,        fill=SF[status], font=SFONTS[status], align='center')
        cs(ws1, r, 7,  row['kaspi'],  align='right', num_fmt='#,##0')
        cs(ws1, r, 8,  row['min'],    align='right', num_fmt='#,##0')
        cs(ws1, r, 9,  row['km'])
        cs(ws1, r, 10, row['brand'])
        ws1.row_dimensions[r].height = 15
        r += 1
 
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
 
# ── Текстовый отчёт ──────────────────────────────────────────────────────────
def extract_date(filename):
    """Извлечь дату из имени файла, например 'Тепловая карта 29.05.xlsx' → '29.05'"""
    import re
    m = re.search(r'(\d{1,2}[.\-]\d{2}(?:[.\-]\d{2,4})?)', filename)
    return m.group(1) if m else datetime.now().strftime('%d.%m')
 
def format_report(rows, filename):
    date_str = extract_date(filename)
 
    # Счётчики
    counts = {}
    for r in rows:
        counts[r['status']] = counts.get(r['status'], 0) + 1
 
    lines = [
        f"📊 Мониторинг цен — {date_str}\n",
        f"🔴 Вне рынка (≥+15%): {counts.get('Вне рынка', 0)} поз.",
        f"🟡 Дороже (0%..+15%): {counts.get('Дороже', 0)} поз.",
        f"⚪ Норма (-5%..0%): {counts.get('Норма', 0)} поз.",
        f"🟢 Дешевле (<-5%): {counts.get('Дешевле', 0)} поз.\n",
    ]
 
    # Позиции «Вне рынка» по сегментам: HARD → SOFT → СЕЗОН
    SEG_ORDER = ['HARD', 'SOFT', 'СЕЗОН']
 
    for seg in SEG_ORDER:
        seg_risks = sorted(
            [r for r in rows if r['segment'] == seg
             and r['status'] == 'Вне рынка' and r.get('pct') and r['name'] != 'АВОКАДО КГ'],
            key=lambda x: -x['pct'])
        if not seg_risks:
            continue
        lines.append(f"— {seg} —")
        for r in seg_risks:
            kaspi_str = f" | Каспий: {r['kaspi']:,}" if r['kaspi'] else ""
            corridor = f"{r['min']:,} – {r['max']:,}" if r.get('max') else f"от {r['min']:,}"
            lines.append(
                f"• {r['name']}\n"
                f"  Магнум: {r['magnum']:,} | Ср.рынок: {r['avg']:,} (коридор: {corridor}){kaspi_str} | +{r['pct']}%\n"
                f"  КМ: {r['km']}"
            )
        lines.append("")
 
    # Сводка по КМ — позиции не в норме (Вне рынка + Дороже)
    km_stats = {}
    for r in rows:
        km = r['km']
        if km == '—':
            continue
        km_stats.setdefault(km, {'вне': 0, 'дороже': 0, 'total': 0})
        km_stats[km]['total'] += 1
        if r['status'] == 'Вне рынка':
            km_stats[km]['вне'] += 1
        elif r['status'] == 'Дороже':
            km_stats[km]['дороже'] += 1
 
    lines.append("👤 По ответственным (не в норме):")
    for km, s in sorted(km_stats.items(), key=lambda x: -(x[1]['вне'] + x[1]['дороже'])):
        not_ok = s['вне'] + s['дороже']
        if not_ok == 0:
            continue
        lines.append(
            f"• {km}: 🔴 {s['вне']} вне рынка, 🟡 {s['дороже']} дороже (из {s['total']} поз.)"
        )
 
    lines.append("")
    lines.append("ℹ️ Ср. цена рынка = Eurospar, Carefood, METRO, Toimart, среднее двух Small (Райымбека + Ауэзова). Оптовка и Рынок Турксиб исключены. Каспий — справочно.")
 
    return "\n".join(lines)
 
# ── Хендлеры ─────────────────────────────────────────────────────────────────
@bot.message_handler(content_types=['document'])
def handle_document(message):
    doc = message.document
    if not doc.file_name.endswith('.xlsx'):
        bot.reply_to(message, "Пришлите файл в формате .xlsx")
        return
 
    bot.reply_to(message, "⏳ Анализирую...")
 
    try:
        file_info = bot.get_file(doc.file_id)
        url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_info.file_path}"
        file_bytes = requests.get(url).content
 
        rows    = run_analysis(file_bytes)
        report  = format_report(rows, doc.file_name)
        excel   = build_excel(rows)
        date_str = datetime.now().strftime('%d.%m')
 
        bot.send_message(message.chat.id, report)
        bot.send_document(message.chat.id, excel,
                          visible_file_name=f"Анализ мониторинга {date_str}.xlsx",
                          caption=f"📎 Полный анализ — {date_str}")
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {e}")
 
@bot.message_handler(func=lambda m: True)
def handle_text(message):
    bot.reply_to(message, "Пришлите файл мониторинга в формате .xlsx 📎")
 
print("✅ Бот запущен (облачный режим)")
bot.infinity_polling()
    bot.reply_to(message, "Пришлите файл мониторинга в формате .xlsx 📎")
 
print("✅ Бот запущен (облачный режим)")
bot.infinity_polling()
